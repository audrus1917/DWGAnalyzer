"""Console entry point for working with DWG/DXF."""

from __future__ import annotations

import asyncio
import json
import logging
import sys

from pathlib import Path

from gettext import gettext as _

from src.parsedwg import constants
from src.parsedwg.explorer import DXFExplorer
from src.parsedwg.utils import (
    print_as_table,
    _save_rows_to_json,
)
from src.parsedwg.utils.args import build_args_parser
from src.parsedwg.commands.parse import handle_parse_command
from src.parsedwg.commands.interpret import handle_interpret_command
from src.parsedwg.constants import ResultRow
from src.parsedwg.ai.docs_ingest import run_documents_ingest


logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s - %(filename)s:%(lineno)d - %(levelname)s - %(message)s",
)
logging.getLogger("ezdxf").disabled = True
logging.getLogger('fontTools.ttLib.ttFont').setLevel(logging.WARNING)
logging.getLogger('matplotlib.font_manager').disabled = True
logger = logging.getLogger(__name__)


def _as_table(rows: list[ResultRow]) -> str:
    if not rows:
        return "No data."

    columns = list(rows[0].keys())
    prepared_rows = [[str(row.get(column, "")) for column in columns] for row in rows]
    widths = {
        column: max(len(column), *(len(values[index]) for values in prepared_rows))
        for index, column in enumerate(columns)
    }

    header = " | ".join(column.ljust(widths[column]) for column in columns)
    separator = "-+-".join("-" * widths[column] for column in columns)
    body = [
        " | ".join(values[index].ljust(widths[column]) for index, column in enumerate(columns))
        for values in prepared_rows
    ]
    return "\n".join([header, separator, *body])


def handle_search_command(
    query: str,
    entity_type: str | None,
    limit: int,
    output_path: Path | None,
    parent_id: str | None = None,
) -> int:
    """Run full-text search against the PostgreSQL entity table."""
    from .db import search_entities

    rows: list[ResultRow] = asyncio.run(search_entities(query, entity_type, limit, parent_id))

    if not rows:
        logger.warning("No results.")
        return constants.OK

    if output_path is not None:
        _save_rows_to_json(output_path, rows)
        logger.info("JSON saved: %s", output_path)
        return constants.OK

    print_as_table(rows)
    return constants.OK


def handle_index_command(
    entity_type: str | None,
    batch_size: int,
    reindex: bool,
) -> int:
    """Generate and save embeddings for entities from the database."""
    from .rag import index_entities

    count = asyncio.run(index_entities(entity_type, batch_size, reindex))
    logger.debug(f"Indexed: {count}")
    return constants.OK


def handle_ask_command(
    question: str,
    entity_type: str | None,
    top_k: int,
    output_path: Path | None,
) -> int:
    """Run a RAG query: vector search plus answer generation through llama."""
    from .rag import ask

    result: ResultRow = asyncio.run(ask(question, entity_type, top_k))

    if output_path is not None:
        _save_rows_to_json(output_path, [result])
        logger.info("JSON saved: %s", output_path)
        return constants.OK

    logger.debug(result["answer"])
    logger.debug("")
    print("Sources:")
    print_as_table(result["sources"])  # type: ignore[arg-type]
    return constants.OK


def handle_process_docs_command(source_path: Path) -> int:
    """Recursively index PDF/DOCX/XLSX/CSV documents into the entity table."""

    summary = run_documents_ingest(source_path)
    print(f"Documents found: {summary['doc_count']}")
    print(f"Entities created in DB: {summary['created_entities']}")
    print(f"Source: {summary['source']}")
    return constants.OK


def handle_export_block_command(
    drawing_path: Path,
    block_name: str,
    output_path: Path | None,
    output_format: str,
) -> int:
    """Export the selected block to the specified format."""

    try:
        explorer = DXFExplorer(drawing_path)
    except FileNotFoundError as exc:
        logger.error("Failed to export block to PNG: %s", exc)
        return constants.ERROR

    if output_format == "png":
        try:
            explorer.export_block_png(block_name, output_path=output_path)
            return constants.OK
        except (FileNotFoundError, RuntimeError, ValueError) as exc:
            logger.error("Failed to export block to PNG: %s", exc)
            return constants.ERROR
    if output_format == "svg":
        try:
            explorer.export_block_svg(block_name, output_path=output_path)
            return constants.OK
        except (FileNotFoundError, RuntimeError, ValueError) as exc:
            logger.error("Failed to export block to SVG: %s", exc)
            return constants.ERROR
    if output_format == "dxf":
        try:
            dxf_text = explorer.export_block_dxf(block_name)
            if output_path is not None:
                resolved_output_path = output_path
                if resolved_output_path.suffix.lower() != ".dxf":
                    resolved_output_path = resolved_output_path.with_suffix(".dxf")
                resolved_output_path.parent.mkdir(parents=True, exist_ok=True)
                resolved_output_path.write_text(dxf_text, encoding="utf-8")
                print(f"DXF saved: {resolved_output_path}")
                return constants.OK

            logger.debug(dxf_text)
            return constants.OK
        except (FileNotFoundError, RuntimeError, ValueError) as exc:
            logger.error("Failed to export block DXF text: %s", exc)
            return constants.ERROR

    logger.error("Unsupported export format: %s", output_format)
    return constants.ERROR


def handle_extract_block_command(
    drawing_path: Path,
    block_name: str,
) -> int:
    """Extract a block into a separate file through DXFExplorer."""

    try:
        explorer = DXFExplorer(drawing_path)
        return int(explorer.extract_block(block_name))
    except (FileNotFoundError, RuntimeError, ValueError) as exc:
        logger.error("Failed to extract block: %s", exc)
        return constants.ERROR


def handle_verify_extraction_command(
    drawing_path: Path,
    file_id: str | None = None,
) -> int:
    """Compare DWG/DXF file contents with what is already stored in the database."""

    from .verify_extraction import format_verification_report, verify_extraction

    try:
        report = asyncio.run(verify_extraction(drawing_path, file_id=file_id))
        print(format_verification_report(report))
        return constants.OK if report["ok"] else constants.ERROR
    except LookupError as e:
        logger.error("File not found in the database: %s", e, exc_info=True)
        return constants.NOT_FOUND
    except (FileNotFoundError, OSError, ValueError) as e:
        logger.error("Verification error: %s", e)
        return constants.ERROR


def _resolve_source_ref_to_drawing_path(source_ref: str, temp_dir: Path) -> Path:
    from .utils import extract_from_zip

    if "::" not in source_ref:
        return Path(source_ref)

    archive_path_str, member_name = source_ref.split("::", 1)
    return extract_from_zip(Path(archive_path_str), member_name, temp_dir)


def _get_block_layout_by_name(doc, block_name: str):
    for layout in doc.layouts:
        if str(layout.name) == block_name:
            return layout
    for block in doc.blocks:
        if str(block.name) == block_name:
            return block
    if block_name == "Model" or block_name.startswith("*Model_Space"):
        return doc.modelspace()
    return None


def _format_point_payload(point: object | None) -> list[float] | None:
    if point is None:
        return None
    x = getattr(point, "x", None)
    y = getattr(point, "y", None)
    z = getattr(point, "z", 0.0)
    if x is None or y is None:
        return None
    return [float(x), float(y), float(z)]


def handle_project_add_command(
    name: str,
    description: str | None,
    created_by: str | None,
) -> int:
    """Create a project."""
    from .db import create_project

    project = asyncio.run(create_project(name=name, description=description, created_by=created_by))
    print(f"Project created: {project['id']}")
    print(f"Name: {project['name']}")
    return constants.OK


def handle_project_update_command(
    project_id: str,
    name: str | None,
    description: str | None,
    created_by: str | None,
) -> int:
    """Update a project."""
    from .db import update_project

    project = asyncio.run(
        update_project(
            project_id=project_id,
            name=name,
            description=description,
            created_by=created_by,
        )
    )
    if project is None:
        print("Project not found.")
        return constants.NOT_FOUND

    print(f"Project updated: {project['id']}")
    print(f"Name: {project['name']}")
    return constants.OK


def handle_project_delete_command(project_id: str, yes: bool) -> int:
    """Delete a project with confirmation."""
    from .db import delete_project

    if not yes:
        answer = input(
            f"Delete project {project_id}? Type YES to confirm: "
        ).strip()
        if answer != "YES":
            print("Deletion cancelled.")
            return constants.ERROR

    deleted = asyncio.run(delete_project(project_id=project_id))
    if not deleted:
        print("Project not found.")
        return constants.NOT_FOUND

    print(f"Project deleted: {project_id}")
    return constants.OK


def handle_category_add_command(
    name: str,
    description: str | None,
    parent_id: str | None,
) -> int:
    """Create a category."""
    from .db import create_category

    category = asyncio.run(
        create_category(name=name, description=description, parent_id=parent_id)
    )
    print(f"Category created: {category['id']}")
    print(f"Name: {category['name']}")
    if category["parent_id"]:
        print(f"Parent: {category['parent_id']}")
    return constants.OK


def handle_category_update_command(
    category_id: str,
    name: str | None,
    description: str | None,
    parent_id: str | None,
) -> int:
    """Update a category."""
    from .db import update_category

    category = asyncio.run(
        update_category(
            category_id=category_id,
            name=name,
            description=description,
            parent_id=parent_id,
        )
    )
    if category is None:
        print("Category not found.")
        return constants.NOT_FOUND

    print(f"Category updated: {category['id']}")
    print(f"Name: {category['name']}")
    if category["parent_id"]:
        print(f"Parent: {category['parent_id']}")
    return constants.OK


def handle_category_delete_command(category_id: str, yes: bool) -> int:
    """Delete a category with confirmation."""
    from .db import delete_category

    if not yes:
        answer = input(
            f"Delete category {category_id}? Type YES to confirm: "
        ).strip()
        if answer != "YES":
            print("Deletion cancelled.")
            return constants.ERROR

    deleted = asyncio.run(delete_category(category_id=category_id))
    if not deleted:
        print("Category not found.")
        return constants.NOT_FOUND

    print(f"Category deleted: {category_id}")
    return constants.OK


def handle_category_list_command(parent_id: str | None) -> int:
    """Show the category list."""
    from .db import list_categories

    rows: list[ResultRow] = []
    for row in asyncio.run(list_categories(parent_id=parent_id)):
        rows.append(
            {
                "id": row["id"],
                "name": row["name"],
                "description": row["description"],
                "parent_id": row["parent_id"],
            }
        )
    if not rows:
        print("No categories.")
        return constants.OK

    print_as_table(rows)
    return constants.OK


def handle_file_stat_from_db_command(
    file_ref: str,
    by_path: bool,
    output_path: Path | None,
) -> int:
    """Collect XLSX statistics for a file already loaded into the database."""

    import uuid as _uuid

    import openpyxl
    import sqlalchemy as sa
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    from .db import async_session_factory
    from .orm import Entity, EntityType

    async def collect_db_stat():
        async with async_session_factory() as session:
            if by_path:
                file_entity = await session.scalar(
                    sa.select(Entity).where(
                        Entity.entity_type == EntityType.FILE,
                        Entity.data["source_ref"].astext == file_ref,
                    )
                )
            else:
                try:
                    file_uuid = _uuid.UUID(file_ref)
                except ValueError:
                    return None, "Invalid file_id UUID"
                file_entity = await session.get(Entity, file_uuid)
            if not file_entity:
                return None, "File entity not found"

            parent_names = []
            parent = file_entity.parent
            while parent:
                parent_names.append(parent.name)
                parent = parent.parent
            parent_dirs = list(reversed(parent_names))

            project = file_entity.project
            project_name = project.name if project else ""

            blocks = await session.execute(
                sa.select(Entity)
                .where(
                    Entity.parent_id == file_entity.id,
                    Entity.entity_type == EntityType.BLOCK,
                )
                .order_by(Entity.name.asc())
            )
            blocks = [block for (block,) in blocks.all()]

            table_blocks = [block for block in blocks if block.is_table]

            primitives = await session.execute(
                sa.select(Entity)
                .where(
                    Entity.file_id == file_entity.id,
                    Entity.entity_type == EntityType.PRIMITIVE,
                )
                .order_by(Entity.id.asc())
            )
            primitives = [primitive for (primitive,) in primitives.all()]

            return {
                "file": file_entity,
                "parent_dirs": parent_dirs,
                "project": project_name,
                "blocks": blocks,
                "table_blocks": table_blocks,
                "primitives": primitives,
            }, None

    stat, err = asyncio.run(collect_db_stat())
    if err:
        print(f"Error: {err}")
        return constants.ERROR
    assert stat is not None

    file_entity = stat["file"]
    parent_dirs = stat["parent_dirs"]
    project = stat["project"]
    blocks = stat["blocks"]
    table_blocks = stat["table_blocks"]
    primitives = stat["primitives"]

    wb = openpyxl.Workbook()

    ws_file = wb.active
    assert ws_file is not None
    ws_file.title = "File"
    ws_file.append(["Parameter", "Value"])
    fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    font = Font(bold=True)
    for cell in ws_file[1]:
        cell.fill = fill
        cell.font = font
    ws_file.append(["File name", file_entity.name])
    ws_file.append(["MD5", file_entity.entity_md5 or ""])
    ws_file.append(["Parent directories", " / ".join(parent_dirs)])
    ws_file.append(["Project", project])
    ws_file.freeze_panes = "A2"
    for row in ws_file.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col_idx, col_cells in enumerate(ws_file.iter_cols(), start=1):
        values = [str(cell.value) for cell in col_cells if cell.value is not None]
        width = max((len(line) for value in values for line in value.splitlines()), default=10)
        ws_file.column_dimensions[get_column_letter(col_idx)].width = min(width + 2, 60)

    ws_blocks = wb.create_sheet("Blocks")
    headers_blocks = ["Name", "Table", "Insert count", "Layers"]
    ws_blocks.append(headers_blocks)
    for cell in ws_blocks[1]:
        cell.fill = fill
        cell.font = font
    ws_blocks.freeze_panes = "A2"
    for block in blocks:
        ws_blocks.append([
            block.name,
            "Yes" if block.is_table else "No",
            "-",
            "-",
        ])
    for row in ws_blocks.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col_idx, col_cells in enumerate(ws_blocks.iter_cols(), start=1):
        values = [str(cell.value) for cell in col_cells if cell.value is not None]
        width = max((len(line) for value in values for line in value.splitlines()), default=10)
        ws_blocks.column_dimensions[get_column_letter(col_idx)].width = min(width + 2, 60)

    ws_tables = wb.create_sheet("Table Blocks")
    ws_tables.append(headers_blocks)
    for cell in ws_tables[1]:
        cell.fill = fill
        cell.font = font
    ws_tables.freeze_panes = "A2"
    for block in table_blocks:
        ws_tables.append([
            block.name,
            "Yes",
            "-",
            "-",
        ])
    for row in ws_tables.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col_idx, col_cells in enumerate(ws_tables.iter_cols(), start=1):
        values = [str(cell.value) for cell in col_cells if cell.value is not None]
        width = max((len(line) for value in values for line in value.splitlines()), default=10)
        ws_tables.column_dimensions[get_column_letter(col_idx)].width = min(width + 2, 60)

    ws_prim = wb.create_sheet("Text Primitives")
    headers_prim = ["Block", "Type", "Text", "Layer", "Location"]
    ws_prim.append(headers_prim)
    for cell in ws_prim[1]:
        cell.fill = fill
        cell.font = font
    ws_prim.freeze_panes = "A2"
    for primitive in primitives:
        ws_prim.append([
            primitive.data.get("block", "") if primitive.data else "",
            primitive.data.get("type", "") if primitive.data else "",
            primitive.data.get("text", "") if primitive.data else "",
            primitive.data.get("layer", "") if primitive.data else "",
            primitive.data.get("location", "") if primitive.data else "",
        ])
    for row in ws_prim.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col_idx, col_cells in enumerate(ws_prim.iter_cols(), start=1):
        values = [str(cell.value) for cell in col_cells if cell.value is not None]
        width = max((len(line) for value in values for line in value.splitlines()), default=10)
        ws_prim.column_dimensions[get_column_letter(col_idx)].width = min(width + 2, 60)

    if output_path is None:
        if by_path:
            stem = Path(file_ref).stem
            output_path = Path(f"{stem}_dbstat.xlsx")
        else:
            output_path = Path(f"{file_ref}_dbstat.xlsx")
    wb.save(output_path)
    print(f"File statistics from the database were saved: {output_path}")
    return constants.OK


def handle_export_blocks_xlsx_command(
    file_ref: str,
    by_path: bool,
    output_path: Path | None,
) -> int:
    """Export a summary table of file blocks from the database to XLSX."""

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    block_rows = _collect_block_export_rows(file_ref=file_ref, by_path=by_path, multiline=True)
    if block_rows is None:
        return constants.ERROR

    if output_path is None:
        if by_path:
            output_path = Path(file_ref).with_suffix(".blocks.xlsx")
        else:
            output_path = Path(f"{file_ref}_blocks.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Blocks"

    headers = [
        "Block name",
        "Related layer names",
        "Full interpretation",
        "Useful attributes",
        "Short interpretation",
        "Block insert count in drawing",
    ]
    ws.append(headers)

    fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in block_rows:
        ws.append([row[column] for column in headers])

    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col_idx, col_cells in enumerate(ws.iter_cols(), start=1):
        values = [str(cell.value) for cell in col_cells if cell.value is not None]
        width = max((len(line) for value in values for line in value.splitlines()), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 2, 60)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Blocks XLSX saved: {output_path}")
    return constants.OK


def handle_export_interpreted_blocks_xlsx_command(output_path: Path | None) -> int:
    """Export all blocks with non-empty short_interpretation to XLSX."""

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    from .db import list_interpreted_blocks_for_export

    try:
        rows = asyncio.run(list_interpreted_blocks_for_export())
    except (LookupError, OSError, RuntimeError, ValueError) as exc:
        logger.error("Failed to collect interpreted blocks: %s", exc)
        return constants.ERROR

    if not rows:
        print("No blocks with non-empty short_interpretation.")
        return constants.OK

    if output_path is None:
        output_path = Path("interpreted_blocks.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Blocks"

    headers = [
        "id",
        "name",
        "description",
        "short_interpretation",
        "full_interpretation",
    ]
    ws.append(headers)

    fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in rows:
        ws.append([row[column] for column in headers])

    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col_idx, col_cells in enumerate(ws.iter_cols(), start=1):
        values = [str(cell.value) for cell in col_cells if cell.value is not None]
        width = max((len(line) for value in values for line in value.splitlines()), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 2, 80)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Interpreted blocks XLSX saved: {output_path}")
    return constants.OK


def _collect_block_export_rows(
    file_ref: str,
    by_path: bool,
    multiline: bool,
) -> list[ResultRow] | None:
    from .db import get_file_id_by_source, list_blocks_for_export

    file_id = file_ref
    if by_path:
        file_id = asyncio.run(get_file_id_by_source(file_ref)) or ""
        if not file_id:
            print("Error: file entity not found")
            return None

    try:
        blocks = asyncio.run(list_blocks_for_export(file_id))
    except (LookupError, OSError, ValueError) as exc:
        logger.error("Failed to collect block data for export: %s", exc)
        return None

    line_separator = "\n" if multiline else "; "
    text_separator = "\n" if multiline else " "
    rows: list[ResultRow] = []
    for block in blocks:
        raw_layers = block.get("layers", [])
        layers = raw_layers if isinstance(raw_layers, list) else []
        layer_names = line_separator.join(
            str(layer.get("name", ""))
            for layer in layers
            if isinstance(layer, dict) and str(layer.get("name", "")).strip()
        )
        attributes = block.get("attributes", {})
        formatted_attributes = (
            line_separator.join(f"{key}: {attributes[key]}" for key in sorted(attributes))
            if isinstance(attributes, dict)
            else ""
        )
        full_interpretation = str(block.get("full_interpretation", "") or "")
        if not multiline:
            full_interpretation = " ".join(full_interpretation.split())
        raw_insert_count = block.get("insert_count", 0)
        insert_count = raw_insert_count if isinstance(raw_insert_count, int) else 0

        rows.append(
            {
                "Block name": str(block.get("name", "") or ""),
                "Related layer names": layer_names,
                "Full interpretation": full_interpretation.replace("\n", text_separator),
                "Useful attributes": formatted_attributes,
                "Short interpretation": str(block.get("short_interpretation", "") or ""),
                "Block insert count in drawing": insert_count,
            }
        )
    return rows


def handle_export_blocks_table_command(
    file_ref: str,
    by_path: bool,
    output_path: Path | None,
) -> int:
    """Print a summary text table of file blocks from the database."""
    block_rows = _collect_block_export_rows(file_ref=file_ref, by_path=by_path, multiline=False)
    if block_rows is None:
        return constants.ERROR
    if not block_rows:
        print("No blocks to export.")
        return constants.OK

    table_text = _as_table(block_rows)
    if output_path is not None:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(table_text + "\n", encoding="utf-8")
        print(f"Block text table saved: {output_path}")
        return constants.OK

    print(table_text)
    return constants.OK


def main(argv: list[str] | None = None) -> int:
    """CLI entry point."""

    parser = build_args_parser()
    argv = list(sys.argv[1:] if argv is None else argv)
    args = parser.parse_args(argv)

    return_code: int = 0
    match args.command:

        case "export":
            return_code = handle_export_block_command(
                drawing_path=Path(args.file_path),
                block_name=args.block_name,
                output_path=Path(args.output) if args.output else None,
                output_format=args.format,
            )

        case "parse":
            if not args.dry and not args.project:
                parser.error(_("The `parse` command requires --project unless --dry is used"))

            return_code = handle_parse_command(
                Path(args.path),
                project_name=args.project,
                dry=args.dry,
                detail_level=args.detail_level,
            )

        case "interpret":
            return_code = handle_interpret_command(
                entity_ids=args.entity_ids,
                entity_type=args.entity_type,
                workers=args.workers,
                dry=args.dry,
            )

        case "verify-extraction":
            return_code = handle_verify_extraction_command(
                drawing_path=Path(args.file_path),
                file_id=args.file_id,
            )

        case "ingest-docs":
            return_code = handle_process_docs_command(Path(args.path))

        case "project-add":
            return_code = handle_project_add_command(
                name=args.name,
                description=args.description,
                created_by=args.created_by,
            )

        case "project-update":
            return_code = handle_project_update_command(
                project_id=args.project_id,
                name=args.name,
                description=args.description,
                created_by=args.created_by,
            )

        case "project-delete":
            return_code = handle_project_delete_command(
                project_id=args.project_id,
                yes=args.yes,
            )

        case "category-add":
            return_code = handle_category_add_command(
                name=args.name,
                description=args.description,
                parent_id=args.parent_id,
            )

        case "category-update":
            return_code = handle_category_update_command(
                category_id=args.category_id,
                name=args.name,
                description=args.description,
                parent_id=args.parent_id,
            )

        case "category-delete":
            return_code = handle_category_delete_command(
                category_id=args.category_id,
                yes=args.yes,
            )

        case "category-list":
            return_code = handle_category_list_command(parent_id=args.parent_id)

        case "export-interpreted-blocks-xlsx":
            return_code = handle_export_interpreted_blocks_xlsx_command(
                output_path=Path(args.output) if args.output else None,
            )

    return return_code

__all__ = ["DXFExplorer", "main"]
