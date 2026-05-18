"""Utilities for inspecting DWG/DXF files and related operations."""

from __future__ import annotations

import hashlib
import logging
import re
from collections.abc import Iterable
from io import StringIO
from typing import Any

from pathlib import Path

from ezdxf.document import Drawing
from ezdxf.filemanagement import new, readfile
from ezdxf.lldxf.tagwriter import TagCollector
from ezdxf.addons.odafc import readfile as read_odafc
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .dxf_analyzer import DrawingAnalyzer
from .table_analysis import TableAnalysis, TextClusterAnalyzer

logger = logging.getLogger(__name__)

type ExplorerRow = dict[str, object]


class DXFExplorer:
    """Inspect DWG/DXF files and perform related operations."""

    def __init__(self, file_path: Path | str):
        """Initialize the DWG/DXF file explorer.

        Raises:
            FileNotFoundError: If file_path does not point to an existing file.
        """
        self.file_path = Path(file_path)
        if not self.file_path.is_file():
            logger.error("File %s was not found.", self.file_path)
            raise FileNotFoundError(f"File {self.file_path} was not found.")

        size_mb = self.file_path.stat().st_size / (1024 * 1024)
        logger.debug("Processing file: %s", self.file_path)
        logger.debug("File size: %.2f MB", size_mb)

    def read_drawing(self) -> Drawing:
        """Read a DWG/DXF file and return the ezdxf Drawing object.

        Returns:
            Loaded Drawing object.
        """

        if self.file_path.suffix.lower() == ".dwg":
            logger.debug(
                "File %s is in DWG format; converting it to DXF first.",
                self.file_path,
            )
            return read_odafc(self.file_path, "ACAD2018")
        else:
            return readfile(self.file_path)

    @staticmethod
    def format_point(point: object | None) -> str:
        """Return a formatted string representation of point coordinates.

        Args:
            point: Point object or tuple-like coordinate value.

        Returns:
            String representation of coordinates or the original value.
        """
        if point is None:
            return "n/a"

        x = getattr(point, "x", None)
        y = getattr(point, "y", None)
        z = getattr(point, "z", 0.0)
        if x is not None and y is not None:
            return f"({x:.2f}, {y:.2f}, {z:.2f})"

        if isinstance(point, (tuple, list)) and len(point) >= 2:
            try:
                x = float(point[0])
                y = float(point[1])
                z = float(point[2]) if len(point) >= 3 else 0.0
            except (TypeError, ValueError):
                return str(point)
            return f"({x:.2f}, {y:.2f}, {z:.2f})"

        return str(point)

    @staticmethod
    def is_point(value: object) -> bool:
        """Return True if the value looks like a coordinate point.

        Supports objects with x/y attributes and tuple-like coordinate containers.
        """
        if hasattr(value, "x") and hasattr(value, "y"):
            return True

        if isinstance(value, (tuple, list)) and len(value) >= 2:
            try:
                float(value[0])
                float(value[1])
            except (TypeError, ValueError):
                return False
            return True

        return False

    @staticmethod
    def get_text_content(entity) -> str:
        """Return the content of an entity text attribute.

        Args:
            entity: DXF entity.

        Returns:
            Text content of the entity or an empty string.
        """

        entity_type = entity.dxftype()
        if entity_type == "TEXT" and entity.dxf.hasattr("text"):
            return entity.dxf.text.strip()

        if entity_type == "MTEXT":
            plain_text = getattr(entity, "plain_text", None)
            if callable(plain_text):
                return str(plain_text()).strip()

        return ""

    @classmethod
    def _collect_entity_layers(
        cls,
        doc,
        entity,
        seen_blocks: set[str] | None = None,
    ) -> set[str]:
        """Collect all layers touched by an entity, including nested blocks.

        Args:
            doc: Loaded ezdxf drawing.
            entity: DXF entity.
            seen_blocks: Already visited blocks used to protect against cycles.

        Returns:
            Set of layer names.
        """
        layers: set[str] = set()
        layer_name = getattr(entity.dxf, "layer", "")
        if layer_name:
            layers.add(str(layer_name))

        if entity.dxftype() != "INSERT" or not entity.dxf.hasattr("name"):
            return layers

        block_name = str(entity.dxf.name)
        if seen_blocks is None:
            seen_blocks = set()
        if block_name in seen_blocks:
            return layers

        block = doc.blocks.get(block_name)
        if block is None:
            return layers

        nested_seen_blocks = {*seen_blocks, block_name}
        for nested_entity in block:
            layers.update(cls._collect_entity_layers(doc, nested_entity, nested_seen_blocks))

        return layers

    @classmethod
    def _get_layout_layers(cls, doc, layout) -> str:
        """Return the comma-separated list of all layout layers.

        Args:
            doc: Loaded ezdxf drawing.
            layout: Layout for which layers should be collected.

        Returns:
            Comma-separated layer string or "-".
        """
        layers: set[str] = set()
        for entity in layout:
            layers.update(cls._collect_entity_layers(doc, entity))
        return ", ".join(sorted(layers)) if layers else "-"

    @classmethod
    def _get_entity_params(cls, entity) -> dict[str, str]:
        """Return a dictionary with entity parameters.

        Args:
            entity: DXF entity.

        Returns:
            Normalized dictionary of entity parameters.
        """
        entity_type = entity.dxftype()
        params: dict[str, str] = {"type": entity_type}

        if entity_type == "INSERT" and entity.dxf.hasattr("name"):
            params["block"] = entity.dxf.name

        for attr_name, value in entity.dxf.all_existing_dxf_attribs().items():
            if cls.is_point(value):
                params[attr_name] = cls.format_point(value)

        text_value = cls.get_text_content(entity)
        if text_value:
            params["text"] = text_value

        if entity_type == "LWPOLYLINE":
            get_points = getattr(entity, "get_points", None)
            if callable(get_points):
                raw_points = get_points("xy")
                if isinstance(raw_points, (list, tuple)):
                    points = [
                        f"({point[0]:.2f}, {point[1]:.2f}, 0.00)"
                        for point in raw_points
                    ]
                    if points:
                        params["points"] = f"[{', '.join(points)}]"
        elif entity_type == "POLYLINE":
            get_points = getattr(entity, "points", None)
            if callable(get_points):
                raw_points = get_points()
                if isinstance(raw_points, Iterable):
                    points = [cls.format_point(point) for point in raw_points]
                    if points:
                        params["points"] = f"[{', '.join(points)}]"
        elif entity_type == "SOLID":
            solid_points = []
            for attr_name in ("vtx0", "vtx1", "vtx2", "vtx3"):
                if entity.dxf.hasattr(attr_name):
                    solid_points.append(cls.format_point(getattr(entity.dxf, attr_name)))
            if solid_points:
                params["points"] = f"[{', '.join(solid_points)}]"

        return params

    @staticmethod
    def _describe_entity(params: dict[str, str]) -> str:
        rendered: list[str] = []
        for key, value in params.items():
            if key in {"block", "text"}:
                rendered.append(f"{key}={value!r}")
            else:
                rendered.append(f"{key}={value}")
        return ", ".join(rendered)

    @classmethod
    def _analyze_text_table(
        cls,
        block,
        x_tolerance: float = 10.0,
        y_tolerance: float = 3.0,
    ) -> TableAnalysis:
        """Analyze a text block as a potential table.

        Args:
            block: DXF block to analyze.
            x_tolerance: Clustering tolerance on X.
            y_tolerance: Clustering tolerance on Y.

        Returns:
            Table analysis result.
        """
        return TextClusterAnalyzer.analyze_table(
            block,
            x_tolerance=x_tolerance,
            y_tolerance=y_tolerance,
        )

    def _export_table_to_xlsx(
        self,
        block_name: str,
        rows: list[list[str]],
        centered_rows: list[int],
        title: str,
        output_dir: Path | None = None,
    ) -> Path:
        """Export a detected block table to XLSX.

        Args:
            block_name: Block name.
            rows: Table rows.
            centered_rows: Row indexes that should be centered and merged.
            title: Table title.
            output_dir: Target output directory.

        Returns:
            Path to the saved XLSX file.
        """
        fallback_name = f"{self.file_path.stem}-{block_name}"
        file_stem = title or fallback_name
        safe_file_name = re.sub(r"[^0-9A-Za-zА-Яа-я._-]+", "_", file_stem).strip("_")
        safe_file_name = safe_file_name or fallback_name
        safe_sheet_name = re.sub(r"[^0-9A-Za-zА-Яа-я._-]+", "_", block_name).strip("_") or "block"
        target_dir = output_dir or self.file_path.parent
        target_dir.mkdir(parents=True, exist_ok=True)
        output_path = target_dir / f"{safe_file_name}.xlsx"

        suffix = 2
        while output_path.exists():
            output_path = target_dir / f"{safe_file_name}_{suffix}.xlsx"
            suffix += 1

        workbook = Workbook()
        sheet = workbook.active
        if sheet is None:
            sheet = workbook.create_sheet()
        sheet.title = safe_sheet_name[:31]
        sheet.freeze_panes = "A2"

        for row in rows:
            sheet.append([value or None for value in row])

        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        if rows:
            last_column = len(rows[0])
            for row_index in centered_rows:
                excel_row = row_index + 1
                if last_column > 1:
                    sheet.merge_cells(
                        start_row=excel_row,
                        start_column=1,
                        end_row=excel_row,
                        end_column=last_column,
                    )
                sheet.cell(row=excel_row, column=1).alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )

        for column_index, column_cells in enumerate(sheet.iter_cols(), start=1):
            values = [str(cell.value) for cell in column_cells if cell.value]
            max_length = max(
                (len(line) for value in values for line in value.splitlines()),
                default=10,
            )
            sheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 2, 60)

        workbook.save(output_path)
        return output_path

    @staticmethod
    def _make_default_export_path(
        drawing: Path,
        file_stem: str,
        suffix: str,
    ) -> Path:
        safe_file_name = re.sub(r"[^0-9A-Za-zА-Яа-я._-]+", "_", file_stem).strip("_")
        safe_file_name = safe_file_name or drawing.stem
        output_path = drawing.parent / f"{safe_file_name}{suffix}"

        duplicate_index = 2
        while output_path.exists():
            output_path = drawing.parent / f"{safe_file_name}_{duplicate_index}{suffix}"
            duplicate_index += 1

        return output_path

    def export_tables_from_db(
        self,
        table_blocks: list[dict[str, object]],
        output_dir: Path,
    ) -> list[Path]:
        """Save XLSX files for table blocks loaded from the database.

        Args:
            table_blocks: Blocks with table data from the database.
            output_dir: Output directory for XLSX files.

        Returns:
            List of created XLSX file paths.
        """
        output_paths: list[Path] = []
        for block_payload in table_blocks:
            block_name = str(block_payload.get("block_name", ""))
            if not block_name:
                continue

            table_payload = block_payload.get("table")
            if not isinstance(table_payload, dict):
                continue

            raw_rows = table_payload.get("rows")
            if not isinstance(raw_rows, list):
                continue

            rows: list[list[str]] = []
            for raw_row in raw_rows:
                if isinstance(raw_row, list):
                    rows.append([str(cell) for cell in raw_row])

            if not rows:
                continue

            title = str(table_payload.get("title") or "")
            output_paths.append(
                self._export_table_to_xlsx(
                    block_name=block_name,
                    rows=rows,
                    centered_rows=[],
                    title=title,
                    output_dir=output_dir,
                )
            )

        return output_paths

    def list_layouts(self) -> list[ExplorerRow]:
        """Return layouts of the current DXF/DWG file.

        Returns:
            List of layouts with names and related layers.
        """

        logger.info("Reading layouts for file: %s", self.file_path)
        doc = self.read_drawing()
        return [
            {
                "drawing": str(self.file_path),
                "layout": layout.name,
                "layers": self._get_layout_layers(doc, layout),
            }
            for layout in doc.layouts
        ]

    def list_blocks(self) -> list[ExplorerRow]:
        """Return blocks of the current DXF/DWG file.

        Returns:
            List of blocks with names and entity counts.
        """

        logger.info("Reading blocks for file: %s", self.file_path)
        doc = self.read_drawing()
        rows: list[ExplorerRow] = []
        for block in doc.blocks:
            logger.debug("Block: %s", block.name)
            rows.append(
                {
                    "drawing": str(self.file_path),
                    "block": block.name,
                    "entity_count": sum(1 for _ in block),
                }
            )
            for entity in block:
                self._get_entity_params(entity)
                # logger.debug("  Entity: %s", self._describe_entity(params))
        return rows

    @staticmethod
    def _is_layout_block_name(block_name: str) -> bool:
        """Return True for internal block names associated with layouts.

        Args:
            block_name: Block name.

        Returns:
            True if the block belongs to internal layout blocks.
        """

        return block_name.startswith("*Model_Space") or block_name.startswith("*Paper_Space")

    @classmethod
    def _collect_block_insert_rows(cls, doc, target_block_name: str) -> list[ExplorerRow]:
        """Return all INSERT entities that reference the given block.

        Args:
            doc: Loaded ezdxf drawing.
            target_block_name: Target block name.

        Returns:
            List of rows describing block usage locations.
        """

        rows: list[ExplorerRow] = []

        for layout in doc.layouts:
            for entity in layout:
                if entity.dxftype() != "INSERT" or not entity.dxf.hasattr("name"):
                    continue
                if str(entity.dxf.name) != target_block_name:
                    continue

                params = cls._get_entity_params(entity)
                rows.append(
                    {
                        "container_type": "layout",
                        "container_name": str(layout.name),
                        **params,
                    }
                )

        for block in doc.blocks:
            parent_block_name = str(block.name)
            if cls._is_layout_block_name(parent_block_name):
                continue

            for entity in block:
                if entity.dxftype() != "INSERT" or not entity.dxf.hasattr("name"):
                    continue
                if str(entity.dxf.name) != target_block_name:
                    continue

                params = cls._get_entity_params(entity)
                rows.append(
                    {
                        "container_type": "block",
                        "container_name": parent_block_name,
                        **params,
                    }
                )

        rows.sort(
            key=lambda row: (
                str(row.get("container_type", "")),
                str(row.get("container_name", "")),
                str(row.get("layer", "")),
                str(row.get("insert", "")),
            )
        )
        return rows

    def describe_block(self, block_name: str) -> dict[str, Any]:
        """Return a JSON-serializable block description from the source file.

        Args:
            block_name: Block name in the file.

        Returns:
            JSON-serializable block description.

        Raises:
            ValueError: If a block with the given block_name is not found.
        """

        logger.info("Collecting block description for '%s' from file: %s", block_name, self.file_path)
        doc = self.read_drawing()
        block = doc.blocks.get(block_name)
        if block is None:
            logger.error("Block '%s' was not found in the file.", block_name)
            raise ValueError(f"Block '{block_name}' was not found in the file.")

        block_description = DrawingAnalyzer.get_short_block_decsription(doc, block_name)
        return block_description

    def list_layer_names(self) -> list[str]:
        """Return layer names from the current DXF/DWG file.

        Returns:
            Sorted list of layer names.
        """

        logger.info("Reading layers for file: %s", self.file_path)
        doc = self.read_drawing()
        return sorted(
            str(layer.dxf.name)
            for layer in doc.layers
            if getattr(layer.dxf, "name", "")
        )

    def extract_block(self, block_name: str) -> int:
        """Extract a block and export a detected table when applicable.

        Args:
            block_name: Block name in the file.

        Returns:
            Operation exit code.

        Raises:
            ValueError: If a block with the given block_name is not found.
        """
        logger.info("Extracting block '%s' from file: %s", block_name, self.file_path)
        doc = self.read_drawing()
        block = doc.blocks.get(block_name)
        if block is None:
            logger.error("Block '%s' was not found in the file.", block_name)
            raise ValueError(f"Block '{block_name}' was not found in the file.")

        for entity in block:
            self._get_entity_params(entity)
            # logger.debug("  Entity: %s", self._describe_entity(params))

        table_stats = self._analyze_text_table(block)
        logger.info(
            "TEXT/MTEXT: total=%s, close on X/Y=%s, X-groups=%s, Y-groups=%s",
            table_stats.total_texts,
            table_stats.table_like_texts,
            len(table_stats.x_clusters),
            len(table_stats.y_clusters),
        )

        if table_stats.is_table:
            output_path = self._export_table_to_xlsx(
                block_name,
                table_stats.rows,
                table_stats.centered_rows,
                table_stats.title,
            )
            logger.info("Table saved to XLSX: %s", output_path)
        else:
            logger.info("Table structure was not detected; XLSX was not created.")
        return 0

    def export_block_png(
        self,
        block_name: str,
        output_path: Path | None = None,
        dpi: int = 300,
    ) -> Path:
        """Export the selected block to PNG.

        Args:
            block_name: Block name.
            output_path: Target path for the PNG.
            dpi: Image resolution.

        Returns:
            Path to the saved PNG.

        Raises:
            ValueError: If the block is not found or export parameters are invalid.
            RuntimeError: If graphical export dependencies are missing.
        """

        return self._export_block_image(
            block_name,
            image_format="png",
            output_path=output_path,
            dpi=dpi,
        )

    def export_block_svg(
        self,
        block_name: str,
        output_path: Path | None = None,
    ) -> Path:
        """Export the selected block to SVG.

        Args:
            block_name: Block name.
            output_path: Target path for the SVG.

        Returns:
            Path to the saved SVG.

        Raises:
            ValueError: If the block is not found or export parameters are invalid.
            RuntimeError: If graphical export dependencies are missing.
        """

        return self._export_block_image(
            block_name,
            image_format="svg",
            output_path=output_path,
        )

    def export_block_dxf(self, block_name: str) -> str:
        """Return DXF text for the selected block.

        Args:
            block_name: Block name.

        Returns:
            DXF text representation of the block.

        Raises:
            ValueError: If a block with the given block_name is not found.
            RuntimeError: If the block cannot be serialized to DXF.
        """

        logger.info("Exporting DXF text for block '%s' from file: %s", block_name, self.file_path)
        doc = self.read_drawing()
        block = doc.blocks.get(block_name)
        if block is None:
            logger.error("Block '%s' was not found in the file.", block_name)
            raise ValueError(f"Block '{block_name}' was not found in the file.")

        if block.block is None or block.endblk is None:
            raise RuntimeError(f"Block '{block_name}' cannot be serialized to DXF.")

        collector = TagCollector(dxfversion=doc.dxfversion)
        block.block.export_dxf(collector)
        for entity in block:
            entity.export_dxf(collector)
        block.endblk.export_dxf(collector)

        stream = StringIO()
        for tag in collector.tags:
            stream.write(tag.dxfstr())
        return stream.getvalue()

    def _export_block_image(
        self,
        block_name: str,
        image_format: str,
        output_path: Path | None = None,
        dpi: int = 300,
    ) -> Path:
        """Export the selected block to an image file.

        Args:
            block_name: Block name.
            image_format: Output file format.
            output_path: Target output path.
            dpi: Image resolution.

        Returns:
            Path to the saved image.

        Raises:
            ValueError: If image_format is unsupported, dpi is invalid, or the block is not found.
            RuntimeError: If matplotlib is missing for export.
        """

        normalized_format = image_format.lower()
        if normalized_format not in {"png", "svg"}:
            raise ValueError(f"Unsupported export format: {image_format}.")

        if dpi <= 0:
            raise ValueError("DPI must be a positive number.")

        logger.info(
            "Exporting block '%s' to %s from file: %s",
            block_name,
            normalized_format.upper(),
            self.file_path,
        )
        doc = self.read_drawing()
        block = doc.blocks.get(block_name)
        if block is None:
            logger.error("Block '%s' was not found in the file.", block_name)
            raise ValueError(f"Block '{block_name}' was not found in the file.")

        try:
            import matplotlib

            matplotlib.use("Agg")

            import matplotlib.pyplot as plt
            from ezdxf.addons.drawing import Frontend, RenderContext
            from ezdxf.addons.drawing.config import Configuration
            from ezdxf.addons.drawing.matplotlib import MatplotlibBackend
            from ezdxf.addons.importer import Importer
        except ImportError as exc:
            raise RuntimeError(
                "matplotlib must be installed to export a block to PNG/SVG."
            ) from exc

        target_doc = new()
        importer = Importer(doc, target_doc)
        importer.import_block(block_name)
        importer.finalize()
        target_doc.modelspace().add_blockref(block_name, (0, 0))

        if output_path is None:
            resolved_output_path = self._make_default_export_path(
                self.file_path,
                f"{self.file_path.stem}-{block_name}",
                f".{normalized_format}",
            )
        else:
            resolved_output_path = output_path
            if resolved_output_path.suffix.lower() != f".{normalized_format}":
                resolved_output_path = resolved_output_path.with_suffix(
                    f".{normalized_format}"
                )
            resolved_output_path.parent.mkdir(parents=True, exist_ok=True)

        figure = plt.figure(figsize=(8, 8))
        try:
            axis = figure.add_axes([0, 0, 1, 1])
            axis.set_axis_off()
            axis.set_aspect("equal", adjustable="datalim")
            axis.margins(0.05)

            context = RenderContext(target_doc)
            backend = MatplotlibBackend(axis)
            Frontend(context, backend, config=Configuration()).draw_layout(target_doc.modelspace())
            axis.autoscale_view()

            figure.savefig(
                resolved_output_path,
                dpi=dpi,
                format=normalized_format,
                bbox_inches="tight",
                pad_inches=0.05,
                facecolor="white",
            )
        finally:
            plt.close(figure)

        logger.info("%s saved: %s", normalized_format.upper(), resolved_output_path)
        return resolved_output_path

    @staticmethod
    def _md5_file(path: Path) -> str:
        h = hashlib.md5()
        with path.open("rb") as f:
            for chunk in iter(lambda: f.read(65536), b""):
                h.update(chunk)
        return h.hexdigest()

    @staticmethod
    def _collect_block_inserts(doc) -> dict[str, dict[str, object]]:
        """Return block_name -> {insert_count, layers} across all layouts."""
        stats: dict[str, dict] = {}
        for layout in doc.layouts:
            for entity in layout:
                if entity.dxftype() != "INSERT" or not entity.dxf.hasattr("name"):
                    continue
                block_name = str(entity.dxf.name)
                layer = str(getattr(entity.dxf, "layer", ""))
                if block_name not in stats:
                    stats[block_name] = {"insert_count": 0, "layers": set()}
                stats[block_name]["insert_count"] += 1
                if layer:
                    stats[block_name]["layers"].add(layer)
        return stats

    @staticmethod
    def _apply_header_style(sheet, row: int) -> None:
        fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
        font = Font(bold=True)
        for cell in sheet[row]:
            cell.fill = fill
            cell.font = font

    @staticmethod
    def _auto_column_widths(sheet) -> None:
        for col_idx, col_cells in enumerate(sheet.iter_cols(), start=1):
            values = [str(c.value) for c in col_cells if c.value is not None]
            width = max((len(line) for v in values for line in v.splitlines()), default=10)
            sheet.column_dimensions[get_column_letter(col_idx)].width = min(width + 2, 60)

    def export_file_stat(self, output_path: Path, project: str = "") -> Path:
        """Collect DXF/DWG file statistics and save them to an XLSX workbook."""
        logger.info("Collecting file statistics: %s", self.file_path)
        doc = self.read_drawing()

        wb = Workbook()
        # --- Sheet 1: File ---
        ws_file = wb.active
        ws_file.title = "File"
        ws_file.append(["Parameter", "Value"])
        self._apply_header_style(ws_file, 1)
        md5 = self._md5_file(self.file_path)
        parent_dirs = [str(p) for p in reversed(self.file_path.parents)]
        ws_file.append(["File name", self.file_path.name])
        ws_file.append(["MD5", md5])
        ws_file.append(["Parent directories", " / ".join(parent_dirs)])
        ws_file.append(["Project", project])
        ws_file.freeze_panes = "A2"
        for row in ws_file.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        self._auto_column_widths(ws_file)

        # --- Sheet 2: Blocks ---
        ws_blocks = wb.create_sheet("Blocks")
        headers_blocks = ["Name", "Table", "Insert count", "Layers"]
        ws_blocks.append(headers_blocks)
        self._apply_header_style(ws_blocks, 1)
        ws_blocks.freeze_panes = "A2"

        insert_stats = self._collect_block_inserts(doc)

        all_block_rows: list[tuple] = []
        for block in doc.blocks:
            block_name = str(block.name)
            table_stats = TextClusterAnalyzer.analyze_table(block)
            is_table = table_stats.is_table
            bstat = insert_stats.get(block_name, {})
            insert_count = bstat.get("insert_count", 0)
            layers = ", ".join(sorted(bstat.get("layers", set())))
            all_block_rows.append((block_name, "Yes" if is_table else "No", insert_count, layers))
            ws_blocks.append([block_name, "Yes" if is_table else "No", insert_count, layers])

        for row in ws_blocks.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        self._auto_column_widths(ws_blocks)

        # --- Sheet 3: Table blocks ---
        ws_tables = wb.create_sheet("Table Blocks")
        ws_tables.append(headers_blocks)
        self._apply_header_style(ws_tables, 1)
        ws_tables.freeze_panes = "A2"

        for row_data in all_block_rows:
            if row_data[1] == "Yes":
                ws_tables.append(list(row_data))

        for row in ws_tables.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        self._auto_column_widths(ws_tables)

        # --- Sheet 4: Text primitives ---
        ws_prim = wb.create_sheet("Text Primitives")
        headers_prim = ["Block", "Type", "Text", "Layer", "Location"]
        ws_prim.append(headers_prim)
        self._apply_header_style(ws_prim, 1)
        ws_prim.freeze_panes = "A2"

        for block in doc.blocks:
            block_name = str(block.name)
            for entity in block:
                entity_type = entity.dxftype()
                if entity_type not in {"TEXT", "MTEXT"}:
                    continue
                text_value = self.get_text_content(entity)
                if not text_value:
                    continue
                layer = str(getattr(entity.dxf, "layer", ""))
                location = ""
                if entity.dxf.hasattr("insert"):
                    location = self.format_point(getattr(entity.dxf, "insert"))
                ws_prim.append([block_name, entity_type, text_value, layer, location])

        for row in ws_prim.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        self._auto_column_widths(ws_prim)

        wb.save(output_path)
        logger.info("Statistics saved: %s", output_path)
        return output_path


__all__ = ["DXFExplorer"]
