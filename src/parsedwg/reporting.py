from __future__ import annotations

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

from .models import ParsedItem

HEADERS = ["Name", "Unit", "Quantity", "Source"]


def _write_headers(sheet, headers: list[str]) -> None:
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column, value=header)
        cell.font = Font(bold=True)
        sheet.column_dimensions[chr(64 + column)].width = 22


def build_workbook_bytes(items: list[ParsedItem]) -> bytes:
    """Build an Excel workbook with specification, work list, estimate, and summary sheets.

    Args:
        items: Items to include in the report.

    Returns:
        Byte content of the finished Excel workbook.

    Raises:
        RuntimeError: If the initial Excel sheet could not be created.
    """
    workbook = Workbook()

    so_sheet = workbook.active
    if so_sheet is None:
        raise RuntimeError("Failed to create the initial Excel sheet.")
    so_sheet.title = "Specification"
    vor_sheet = workbook.create_sheet("Work List")
    estimate_sheet = workbook.create_sheet("Estimate")
    summary_sheet = workbook.create_sheet("Summary")

    _write_headers(so_sheet, HEADERS)
    _write_headers(vor_sheet, HEADERS)
    _write_headers(estimate_sheet, [*HEADERS[:3], "Unit Price", "Total", "Source"])

    so_row = 2
    vor_row = 2
    estimate_row = 2

    for item in items:
        if item.section in {"equipment", "materials"}:
            so_sheet.append([item.name, item.unit, item.quantity, item.source])
            so_row += 1
        if item.section == "works":
            vor_sheet.append([item.name, item.unit, item.quantity, item.source])
            vor_row += 1

        estimate_sheet.cell(row=estimate_row, column=1, value=item.name)
        estimate_sheet.cell(row=estimate_row, column=2, value=item.unit)
        estimate_sheet.cell(row=estimate_row, column=3, value=item.quantity)
        estimate_sheet.cell(row=estimate_row, column=4, value=item.unit_price)
        estimate_sheet.cell(row=estimate_row, column=5, value=f"=C{estimate_row}*D{estimate_row}")
        estimate_sheet.cell(row=estimate_row, column=6, value=item.source)
        estimate_row += 1

    summary_sheet["A1"] = "Generated at"
    summary_sheet["B1"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    summary_sheet["A2"] = "Total items"
    summary_sheet["B2"] = len(items)
    summary_sheet["A3"] = "Equipment and materials"
    summary_sheet["B3"] = sum(1 for item in items if item.section in {"equipment", "materials"})
    summary_sheet["A4"] = "Works"
    summary_sheet["B4"] = sum(1 for item in items if item.section == "works")

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
