from pathlib import Path

from openpyxl import load_workbook

from parsedwg.cli import handle_export_interpreted_blocks_xlsx_command
from parsedwg.utils.args import build_args_parser


def test_build_args_parser_accepts_export_interpreted_blocks_xlsx_command() -> None:
    parser = build_args_parser()

    args = parser.parse_args(["export-interpreted-blocks-xlsx", "-o", "blocks.xlsx"])

    assert args.command == "export-interpreted-blocks-xlsx"
    assert args.output == "blocks.xlsx"


def test_handle_export_interpreted_blocks_xlsx_command_writes_expected_columns(
    tmp_path: Path,
    monkeypatch,
) -> None:
    async def fake_list_interpreted_blocks_for_export() -> list[dict[str, str]]:
        return [
            {
                "id": "101",
                "name": "BLOCK_A",
                "description": "Описание A",
                "short_interpretation": "Краткая A",
                "full_interpretation": "Полная A",
            },
            {
                "id": "102",
                "name": "BLOCK_B",
                "description": "Описание B",
                "short_interpretation": "Краткая B",
                "full_interpretation": "Полная B",
            },
        ]

    monkeypatch.setattr(
        "parsedwg.db.list_interpreted_blocks_for_export",
        fake_list_interpreted_blocks_for_export,
    )

    output_path = tmp_path / "interpreted_blocks.xlsx"
    exit_code = handle_export_interpreted_blocks_xlsx_command(output_path)

    assert exit_code == 0
    assert output_path.exists()

    workbook = load_workbook(output_path)
    sheet = workbook.active
    assert sheet is not None

    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == (
        "id",
        "name",
        "description",
        "short_interpretation",
        "full_interpretation",
    )
    assert rows[1] == ("101", "BLOCK_A", "Описание A", "Краткая A", "Полная A")
    assert rows[2] == ("102", "BLOCK_B", "Описание B", "Краткая B", "Полная B")