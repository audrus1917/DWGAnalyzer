"""Тесты для команды file-stat."""

from pathlib import Path

import openpyxl
from ezdxf.filemanagement import new

from parsedwg.cli import main


def _make_dxf(tmp_path: Path) -> Path:
    source_path = tmp_path / "sample.dxf"
    doc = new()

    block_a = doc.blocks.new("BLOCK_A")
    block_a.add_text("Привет мир", dxfattribs={"layer": "LAYER_1", "insert": (0, 0, 0)})
    block_a.add_text("Строка два", dxfattribs={"layer": "LAYER_1", "insert": (0, 10, 0)})

    doc.blocks.new("BLOCK_B")

    ms = doc.modelspace()
    ms.add_blockref("BLOCK_A", (0, 0), dxfattribs={"layer": "LAYER_2"})
    ms.add_blockref("BLOCK_A", (10, 0), dxfattribs={"layer": "LAYER_2"})

    doc.saveas(source_path)
    return source_path


def test_file_stat_creates_xlsx_with_four_sheets(tmp_path) -> None:
    source_path = _make_dxf(tmp_path)
    output_path = tmp_path / "stat.xlsx"

    exit_code = main(["file-stat", str(source_path), "-o", str(output_path)])

    assert exit_code == 0
    assert output_path.exists()

    wb = openpyxl.load_workbook(output_path)
    assert set(wb.sheetnames) == {"Файл", "Блоки", "Блоки-таблицы", "Текстовые примитивы"}


def test_file_stat_file_sheet_contains_name_and_md5(tmp_path) -> None:
    source_path = _make_dxf(tmp_path)
    output_path = tmp_path / "stat.xlsx"

    main(["file-stat", str(source_path), "-o", str(output_path)])

    wb = openpyxl.load_workbook(output_path)
    ws = wb["Файл"]
    values = {row[0].value: row[1].value for row in ws.iter_rows(min_row=2)}

    assert values["Имя файла"] == "sample.dxf"
    assert len(str(values["MD5"] or "")) == 32


def test_file_stat_project_appears_in_file_sheet(tmp_path) -> None:
    source_path = _make_dxf(tmp_path)
    output_path = tmp_path / "stat.xlsx"

    main(["file-stat", str(source_path), "-o", str(output_path), "--project", "Башня А"])

    wb = openpyxl.load_workbook(output_path)
    ws = wb["Файл"]
    values = {row[0].value: row[1].value for row in ws.iter_rows(min_row=2)}

    assert values["Проект"] == "Башня А"


def test_file_stat_blocks_sheet_has_block_a(tmp_path) -> None:
    source_path = _make_dxf(tmp_path)
    output_path = tmp_path / "stat.xlsx"

    main(["file-stat", str(source_path), "-o", str(output_path)])

    wb = openpyxl.load_workbook(output_path)
    ws = wb["Блоки"]
    headers = [cell.value for cell in ws[1]]
    assert headers == ["Наименование", "Таблица", "Добавлен (раз)", "Слои"]

    rows = {row[0]: row for row in ws.iter_rows(min_row=2, values_only=True)}
    assert "BLOCK_A" in rows
    block_a_row = rows["BLOCK_A"]
    assert block_a_row[2] == 2  # вставлен дважды
    assert "LAYER_2" in str(block_a_row[3] or "")


def test_file_stat_text_primitives_sheet_has_texts(tmp_path) -> None:
    source_path = _make_dxf(tmp_path)
    output_path = tmp_path / "stat.xlsx"

    main(["file-stat", str(source_path), "-o", str(output_path)])

    wb = openpyxl.load_workbook(output_path)
    ws = wb["Текстовые примитивы"]
    headers = [cell.value for cell in ws[1]]
    assert headers == ["Блок", "Тип", "Текст", "Слой", "Локация"]

    texts = [row[2] for row in ws.iter_rows(min_row=2, values_only=True)]
    assert "Привет мир" in texts
    assert "Строка два" in texts


def test_file_stat_default_output_path(tmp_path) -> None:
    source_path = _make_dxf(tmp_path)

    exit_code = main(["file-stat", str(source_path)])

    assert exit_code == 0
    default_output = source_path.with_suffix(".xlsx")
    assert default_output.exists()


def test_file_stat_exports_db_tables_to_same_directory(tmp_path, monkeypatch) -> None:
    source_path = _make_dxf(tmp_path)
    output_path = tmp_path / "stat.xlsx"

    async def fake_get_table_blocks_for_source(_source_ref: str):
        return [
            {
                "block_name": "DB_BLOCK",
                "table": {
                    "title": "DB_TABLE_TITLE",
                    "rows": [["H1", "H2"], ["A", "B"]],
                },
            }
        ]

    monkeypatch.setattr(
        "parsedwg.db.get_table_blocks_for_source",
        fake_get_table_blocks_for_source,
    )

    exit_code = main(["file-stat", str(source_path), "-o", str(output_path), "--db-tables"])

    assert exit_code == 0
    assert output_path.exists()
    db_table_xlsx = tmp_path / "DB_TABLE_TITLE.xlsx"
    assert db_table_xlsx.exists()


def test_file_stat_exports_db_tables_by_id(tmp_path, monkeypatch) -> None:
    source_path = _make_dxf(tmp_path)
    output_path = tmp_path / "stat.xlsx"

    async def fake_get_file_id_by_source(_source_ref: str):
        return "123e4567-e89b-12d3-a456-426614174000"

    async def fake_get_table_blocks_by_file_id(_file_id: str):
        return [
            {
                "block_name": "DB_BLOCK_ID",
                "table": {
                    "title": "DB_TABLE_TITLE_ID",
                    "rows": [["H1", "H2"], ["A", "B"]],
                },
            }
        ]

    monkeypatch.setattr(
        "parsedwg.db.get_file_id_by_source",
        fake_get_file_id_by_source,
    )
    monkeypatch.setattr(
        "parsedwg.db.get_table_blocks_by_file_id",
        fake_get_table_blocks_by_file_id,
    )

    exit_code = main(["file-stat", str(source_path), "-o", str(output_path), "--db-tables-by-id"])

    assert exit_code == 0
    assert output_path.exists()
    db_table_xlsx = tmp_path / "DB_TABLE_TITLE_ID.xlsx"
    assert db_table_xlsx.exists()


def test_export_blocks_xlsx_by_path_creates_expected_sheet(tmp_path, monkeypatch) -> None:
    source_path = _make_dxf(tmp_path)
    output_path = tmp_path / "blocks.xlsx"

    async def fake_get_file_id_by_source(_source_ref: str):
        return "123e4567-e89b-12d3-a456-426614174000"

    async def fake_list_blocks_for_export(_file_id: str):
        return [
            {
                "name": "SOCKET_220V",
                "layers": [
                    {"name": "Electrical", "short_interpretation": "Электрика"},
                    {"name": "Sockets", "short_interpretation": None},
                ],
                "full_interpretation": "Розетка для подключения\nэлектроприборов",
                "attributes": {"power": "220V", "phase": "1"},
                "short_interpretation": "Розетка 220V",
                "insert_count": 50,
            }
        ]

    monkeypatch.setattr("parsedwg.db.get_file_id_by_source", fake_get_file_id_by_source)
    monkeypatch.setattr("parsedwg.db.list_blocks_for_export", fake_list_blocks_for_export)

    exit_code = main([
        "export-blocks-xlsx",
        str(source_path),
        "--by-path",
        "-o",
        str(output_path),
    ])

    assert exit_code == 0
    assert output_path.exists()

    wb = openpyxl.load_workbook(output_path)
    ws = wb["Блоки"]
    headers = [cell.value for cell in ws[1]]
    assert headers == [
        "Название блока",
        "Названия связанных слоев",
        "Интерпретация полная",
        "Полезные атрибуты",
        "Интерпретация краткая",
        "Количество вхождений блока в чертеж",
    ]

    row = [cell.value for cell in ws[2]]
    assert row[0] == "SOCKET_220V"
    assert row[1] == "Electrical\nSockets"
    assert row[2] == "Розетка для подключения\nэлектроприборов"
    assert row[3] == "phase: 1\npower: 220V"
    assert row[4] == "Розетка 220V"
    assert row[5] == 50


def test_export_blocks_table_by_path_prints_expected_ascii(tmp_path, monkeypatch, capsys) -> None:
    source_path = _make_dxf(tmp_path)

    async def fake_get_file_id_by_source(_source_ref: str):
        return "123e4567-e89b-12d3-a456-426614174000"

    async def fake_list_blocks_for_export(_file_id: str):
        return [
            {
                "name": "SOCKET_220V",
                "layers": [
                    {"name": "Electrical", "short_interpretation": "Электрика"},
                    {"name": "Sockets", "short_interpretation": None},
                ],
                "full_interpretation": "Розетка для подключения\nэлектроприборов",
                "attributes": {"power": "220V", "phase": "1"},
                "short_interpretation": "Розетка 220V",
                "insert_count": 50,
            }
        ]

    monkeypatch.setattr("parsedwg.db.get_file_id_by_source", fake_get_file_id_by_source)
    monkeypatch.setattr("parsedwg.db.list_blocks_for_export", fake_list_blocks_for_export)

    exit_code = main([
        "export-blocks-table",
        str(source_path),
        "--by-path",
    ])

    assert exit_code == 0
    captured = capsys.readouterr()
    assert "Название блока" in captured.out
    assert "SOCKET_220V" in captured.out
    assert "Electrical; Sockets" in captured.out
    assert "phase: 1; power: 220V" in captured.out
    assert "Розетка 220V" in captured.out
    assert "50" in captured.out
