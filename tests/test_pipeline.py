from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory

import ezdxf
from openpyxl import load_workbook

from parsedwg.dxf_text_copy import copy_text_entities
from parsedwg.models import ParsedItem
from parsedwg.parsers import parse_item_line
from parsedwg.reporting import build_workbook_bytes
from parsedwg.service import summarize_items


def test_parse_item_line_extracts_quantity_and_unit():
    item = parse_item_line("Кабель ВВГнг 3х1,5 - 120 м", source="drawing")

    assert item is not None
    assert item.name == "Кабель ВВГнг 3х1,5"
    assert item.quantity == 120
    assert item.unit == "м"
    assert item.section == "materials"


def test_summarize_items_groups_duplicates():
    items = [
        ParsedItem(name="Светильник LED", quantity=4, unit="шт", section="equipment", source="dwg"),
        ParsedItem(name="Светильник LED", quantity=2, unit="шт", section="equipment", source="note"),
    ]

    grouped = summarize_items(items)

    assert len(grouped) == 1
    assert grouped[0].quantity == 6
    assert grouped[0].source == "dwg, note"


def test_build_workbook_contains_required_sheets():
    items = [
        ParsedItem(name="Светильник LED", quantity=6, unit="шт", section="equipment", source="dwg"),
        ParsedItem(name="Монтаж кабеля", quantity=120, unit="м", section="works", source="dwg"),
    ]

    payload = build_workbook_bytes(items)
    workbook = load_workbook(BytesIO(payload))

    assert workbook.sheetnames == ["СО", "ВОР", "Смета", "Сводка"]
    assert workbook["СО"]["A2"].value == "Светильник LED"
    assert workbook["ВОР"]["A2"].value == "Монтаж кабеля"


def test_copy_text_entities_filters_non_text_entities():
    with TemporaryDirectory() as temp_dir:
        temp_path = Path(temp_dir)
        source_path = temp_path / "source.dxf"
        target_path = temp_path / "target.dxf"

        source_doc = ezdxf.new()
        source_doc.layers.add("NOTES", color=3)
        source_doc.styles.add("ANNOTATION", font="Arial.ttf")
        source_msp = source_doc.modelspace()
        source_msp.add_line((0, 0), (10, 10))
        source_msp.add_circle((5, 5), radius=2)
        source_msp.add_text("Марка 1", dxfattribs={"layer": "NOTES", "style": "ANNOTATION"})
        source_msp.add_mtext("Примечание")
        sheet_layout = source_doc.layout("Layout1")
        sheet_layout.add_text("Текст на листе")
        source_doc.saveas(source_path)

        copied_count = copy_text_entities(source_path, target_path)

        assert copied_count == 3

        target_doc = ezdxf.readfile(target_path)
        entity_types = [entity.dxftype() for entity in target_doc.modelspace()]
        layout_entity_types = [entity.dxftype() for entity in target_doc.layout("Layout1")]

        assert entity_types == ["TEXT", "MTEXT"]
        assert layout_entity_types == ["TEXT"]
        assert "NOTES" in target_doc.layers
        assert "ANNOTATION" in target_doc.styles


def test_copy_text_entities_modelspace_only_skips_layout_text():
    with TemporaryDirectory() as temp_dir:
        temp_path = Path(temp_dir)
        source_path = temp_path / "source.dxf"
        target_path = temp_path / "target_modelspace.dxf"

        source_doc = ezdxf.new()
        source_doc.modelspace().add_text("Текст модели")
        source_doc.layout("Layout1").add_text("Текст листа")
        source_doc.saveas(source_path)

        copied_count = copy_text_entities(source_path, target_path, modelspace_only=True)

        assert copied_count == 1

        target_doc = ezdxf.readfile(target_path)
        modelspace_texts = [entity.dxf.text for entity in target_doc.modelspace()]
        layout_entities = list(target_doc.layout("Layout1"))

        assert modelspace_texts == ["Текст модели"]
        assert layout_entities == []
