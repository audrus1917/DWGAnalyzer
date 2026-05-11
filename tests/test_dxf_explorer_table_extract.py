from pathlib import Path
from tempfile import TemporaryDirectory

from ezdxf.filemanagement import new
from openpyxl import load_workbook

from parsedwg.cli import DXFExplorer


def test_extract_block_exports_table_to_xlsx() -> None:
    with TemporaryDirectory() as temp_dir:
        source_path = Path(temp_dir) / "table.dxf"

        doc = new()
        block = doc.blocks.new("TABLE_A")
        block.add_text("H1", dxfattribs={"insert": (0, 20, 0)})
        block.add_text("H2", dxfattribs={"insert": (50, 20, 0)})
        block.add_text("R1C1", dxfattribs={"insert": (0, 10, 0)})
        block.add_text("R1C2", dxfattribs={"insert": (50, 10, 0)})
        block.add_mtext("R2C1").set_location((0, 0, 0))
        block.add_mtext("R2C2").set_location((50, 0, 0))
        doc.saveas(source_path)

        explorer = DXFExplorer(source_path)
        explorer.extract_block("TABLE_A")

        workbook_path = Path(temp_dir) / "table-TABLE_A.xlsx"
        workbook = load_workbook(workbook_path)
        sheet = workbook.active
        assert sheet is not None

        rows = list(sheet.iter_rows(values_only=True))
        assert rows[0][:2] == ("H1", "H2")
        assert rows[1][:2] == ("R1C1", "R1C2")
        assert rows[2][:2] == ("R2C1", "R2C2")


def test_extract_block_exports_centered_headers_into_same_columns() -> None:
    with TemporaryDirectory() as temp_dir:
        source_path = Path(temp_dir) / "centered-header-table.dxf"

        doc = new()
        block = doc.blocks.new("CENTERED_TABLE")
        block.add_text("№", dxfattribs={"insert": (10, 30, 0)})
        block.add_text("Описание", dxfattribs={"insert": (80, 30, 0)})
        block.add_text("Ед.", dxfattribs={"insert": (150, 30, 0)})
        block.add_text("1", dxfattribs={"insert": (0, 20, 0)})
        block.add_text("Насос", dxfattribs={"insert": (40, 20, 0)})
        block.add_text("шт", dxfattribs={"insert": (140, 20, 0)})
        block.add_text("2", dxfattribs={"insert": (0, 10, 0)})
        block.add_mtext("Клапан").set_location((40, 10, 0))
        block.add_text("компл", dxfattribs={"insert": (140, 10, 0)})
        doc.saveas(source_path)

        explorer = DXFExplorer(source_path)
        explorer.extract_block("CENTERED_TABLE")

        workbook_path = Path(temp_dir) / "centered-header-table-CENTERED_TABLE.xlsx"
        workbook = load_workbook(workbook_path)
        sheet = workbook.active
        assert sheet is not None

        rows = list(sheet.iter_rows(values_only=True))
        assert rows[0][:3] == ("№", "Описание", "Ед.")
        assert rows[1][:3] == ("1", "Насос", "шт")
        assert rows[2][:3] == ("2", "Клапан", "компл")


def test_extract_block_exports_title_and_full_width_section_rows() -> None:
    with TemporaryDirectory() as temp_dir:
        source_path = Path(temp_dir) / "sectioned-table.dxf"

        doc = new()
        block = doc.blocks.new("SECTION_TABLE")
        block.add_text("Main Title", dxfattribs={"insert": (75, 40, 0)})
        block.add_text("№", dxfattribs={"insert": (10, 30, 0)})
        block.add_text("Описание", dxfattribs={"insert": (80, 30, 0)})
        block.add_text("Ед.", dxfattribs={"insert": (150, 30, 0)})
        block.add_text("1", dxfattribs={"insert": (0, 20, 0)})
        block.add_text("Насос", dxfattribs={"insert": (40, 20, 0)})
        block.add_text("шт", dxfattribs={"insert": (140, 20, 0)})
        block.add_text("Раздел 1", dxfattribs={"insert": (75, 10, 0)})
        block.add_text("2", dxfattribs={"insert": (0, 0, 0)})
        block.add_text("Клапан", dxfattribs={"insert": (40, 0, 0)})
        block.add_text("компл", dxfattribs={"insert": (140, 0, 0)})
        doc.saveas(source_path)

        explorer = DXFExplorer(source_path)
        explorer.extract_block("SECTION_TABLE")

        workbook_path = Path(temp_dir) / "Main_Title.xlsx"
        workbook = load_workbook(workbook_path)
        sheet = workbook.active
        assert sheet is not None

        rows = list(sheet.iter_rows(values_only=True))
        assert rows[0][:3] == ("№", "Описание", "Ед.")
        assert rows[1][:3] == ("1", "Насос", "шт")
        assert rows[2][:3] == ("2", "Клапан", "компл")
        assert list(sheet.merged_cells.ranges) == []


def test_extract_block_keeps_name_out_of_number_column() -> None:
    with TemporaryDirectory() as temp_dir:
        source_path = Path(temp_dir) / "name-header-table.dxf"

        doc = new()
        block = doc.blocks.new("HEADER_TABLE")
        block.add_text("№ п/п", dxfattribs={"insert": (5, 20, 0)})
        block.add_text("Наименование", dxfattribs={"insert": (15, 20, 0)})
        block.add_text("Ед.", dxfattribs={"insert": (140, 20, 0)})
        block.add_text("1", dxfattribs={"insert": (0, 10, 0)})
        block.add_text("Насос", dxfattribs={"insert": (100, 10, 0)})
        block.add_text("шт", dxfattribs={"insert": (140, 10, 0)})
        doc.saveas(source_path)

        explorer = DXFExplorer(source_path)
        explorer.extract_block("HEADER_TABLE")

        workbook_path = Path(temp_dir) / "name-header-table-HEADER_TABLE.xlsx"
        workbook = load_workbook(workbook_path)
        sheet = workbook.active
        assert sheet is not None

        rows = list(sheet.iter_rows(values_only=True))
        assert rows[0][:3] == ("№ п/п", "Наименование", "Ед.")
        assert rows[1][:3] == ("1", "Насос", "шт")


def test_extract_block_skips_consecutive_centered_section_rows() -> None:
    with TemporaryDirectory() as temp_dir:
        source_path = Path(temp_dir) / "stacked-sections.dxf"

        doc = new()
        block = doc.blocks.new("STACKED_SECTIONS")
        block.add_text("№", dxfattribs={"insert": (10, 30, 0)})
        block.add_text("Наименование", dxfattribs={"insert": (80, 30, 0)})
        block.add_text("Ед.", dxfattribs={"insert": (150, 30, 0)})
        block.add_text("Раздел 1", dxfattribs={"insert": (75, 20, 0)})
        block.add_text("Подраздел А", dxfattribs={"insert": (75, 10, 0)})
        block.add_text("1", dxfattribs={"insert": (0, 0, 0)})
        block.add_text("Насос", dxfattribs={"insert": (40, 0, 0)})
        block.add_text("шт", dxfattribs={"insert": (140, 0, 0)})
        doc.saveas(source_path)

        explorer = DXFExplorer(source_path)
        explorer.extract_block("STACKED_SECTIONS")

        workbook_path = Path(temp_dir) / "stacked-sections-STACKED_SECTIONS.xlsx"
        workbook = load_workbook(workbook_path)
        sheet = workbook.active
        assert sheet is not None

        rows = list(sheet.iter_rows(values_only=True))
        assert rows[0][:3] == ("№", "Наименование", "Ед.")
        assert rows[1][:3] == ("1", "Насос", "шт")
        assert len(rows) == 2
        assert list(sheet.merged_cells.ranges) == []


def test_extract_block_skips_rows_with_missing_columns() -> None:
    with TemporaryDirectory() as temp_dir:
        source_path = Path(temp_dir) / "missing-columns.dxf"

        doc = new()
        block = doc.blocks.new("MISSING_COLUMNS")
        block.add_text("№", dxfattribs={"insert": (10, 30, 0)})
        block.add_text("Наименование", dxfattribs={"insert": (80, 30, 0)})
        block.add_text("Ед.", dxfattribs={"insert": (150, 30, 0)})
        block.add_text("Подраздел", dxfattribs={"insert": (40, 20, 0)})
        block.add_text("А", dxfattribs={"insert": (140, 20, 0)})
        block.add_text("1", dxfattribs={"insert": (0, 10, 0)})
        block.add_text("Насос", dxfattribs={"insert": (40, 10, 0)})
        block.add_text("шт", dxfattribs={"insert": (140, 10, 0)})
        doc.saveas(source_path)

        explorer = DXFExplorer(source_path)
        explorer.extract_block("MISSING_COLUMNS")

        workbook_path = Path(temp_dir) / "missing-columns-MISSING_COLUMNS.xlsx"
        workbook = load_workbook(workbook_path)
        sheet = workbook.active
        assert sheet is not None

        rows = list(sheet.iter_rows(values_only=True))
        assert rows[0][:3] == ("№", "Наименование", "Ед.")
        assert rows[1][:3] == ("1", "Насос", "шт")
        assert len(rows) == 2


def test_analyze_text_table_uses_top_y_for_row_clusters() -> None:
    doc = new()
    block = doc.blocks.new("Y_TOP_TABLE")
    block.add_text("H1", dxfattribs={"insert": (0, 20, 0)})
    block.add_text("H2", dxfattribs={"insert": (50, 18, 0)})
    block.add_text("R1C1", dxfattribs={"insert": (0, 10, 0)})
    block.add_text("R1C2", dxfattribs={"insert": (50, 8, 0)})

    analysis = DXFExplorer._analyze_text_table(block, y_tolerance=3.0)

    assert [cluster.center for cluster in analysis.y_clusters] == [20.0, 10.0]